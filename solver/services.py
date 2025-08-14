import re
import unicodedata
from math import ceil
from ortools.constraint_solver import pywrapcp, routing_enums_pb2
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO

def get_matrices(coords, annotations=("distance","duration")):
    """
    Interroge OSRM pour obtenir les matrices des distances & durées.
    Retourne le tuple (distMat, durMat) ou lève une exception.
    """
    # Communication avec l'API
    base = "https://router.project-osrm.org/table/v1/driving/" # URL initial sans paramètre
    url  = base + coords # URL final avec ajout des paramètres
    resp = requests.get(url, params={"annotations": ",".join(annotations)}) # Appel à OSRM
    resp.raise_for_status() # Résultat de l'appel
    data = resp.json() # Récuperation du resultat

    # Levée d'exception
    if "distances" not in data or "durations" not in data:
        raise ValueError("OSRM n'a pas renvoyé les matrices attendues.")
    
    # Création des matrices de distances et de durées renvoyées comme résultat
    distMat = [[int(d) for d in row] for row in data["distances"]]
    durMat  = [[int(d) for d in row] for row in data["durations"]]
    return distMat, durMat

def format_solution(manager, routing, solution, distMat, durMat, groupe, etablissement,
                    virtual_start, virtual_end):
    """
    Retourne un dict sérialisable incluant pour chaque nœud:
      - vehicle_id
      - sequence de prénoms+nom
      - distance et durée totales

    Si virtual_start/end est True, les nœuds virtuels sont ignorés dans l'affichage.
    """
    # Pour débuger
    print("STARTS:", [manager.IndexToNode(routing.Start(vid)) for vid in range(routing.vehicles())])
    print("ENDS:", [manager.IndexToNode(routing.End(vid)) for vid in range(routing.vehicles())])
    print("Virtual start:", virtual_start, "Virtual end:", virtual_end)
    
    # Initialisation de variables
    routes = []
    enfants = list(groupe.enfants.all())

    # Index de l'établissement (dépend de la présence du dépôt virtuel)
    depot_node = 0 if not(virtual_start or virtual_end) else 1

    for vid in range(routing.vehicles()): # On boucle pour chaque véhicule
        index = routing.Start(vid) # Donne l'indice OR-Tools du point de départ du véhicule vid
        seq, noms, prenoms, adresses, villes, distances, durees, dist_tot, dur_tot, = [], [], [], [], [], [], [], 0, 0 # seq = séquence de noeuds visités | dist = distance totale de la tournée | dur = durée totale de la tournée

        while not routing.IsEnd(index): # On regarde si on est au dernier noeud de la tournée
            node = manager.IndexToNode(index) # conversion de l'index OR-Tools en indice réel

            # On regarde si le noeud sur lequel nous sommes est un noeud virtuel
            is_virtual_start = virtual_start and node == 0

            if not is_virtual_start:
                if node == depot_node:
                    seq.append("Établissement")
                    prenoms.append(etablissement.code)
                    noms.append(etablissement.nom)
                    adresses.append(etablissement.adresse.num_et_rue)
                    villes.append(etablissement.adresse.ville)
                else:
                    if virtual_start or virtual_end:
                        enfant_index = node - 2
                    else:
                        enfant_index = node - 1 # L'ordre se réfère aux ajouts dans coord_list, ainsi il s'agit de regarder par rapport à ça pour connaitre l'index des enfants
                    if 0 <= enfant_index < len(enfants): # Mesure de sécurité, évite un index out of range si erreur d'offset ou de calcul
                        child = enfants[enfant_index]
                        seq.append(f"{child.prenom} {child.nom}")
                        noms.append(child.nom)
                        prenoms.append(child.prenom)
                        adresses.append(child.adresse.num_et_rue)
                        villes.append(child.adresse.ville)

                    else:
                        seq.append(f"[Nœud {node}]") 
                        adresses.append("None") # Sécurité

            nxt = solution.Value(routing.NextVar(index)) # On passe au noeud suivant
            nnode = manager.IndexToNode(nxt) # On cherche l'indice réel correspondant
            
            # Récuperation du coût et de la durée de l'étape
            dist = distMat[node][nnode]
            dur  = durMat[node][nnode]
            
            # Incrémentation des coûts totaux
            dist_tot += dist
            dur_tot  += dur

            # Ajout des coûts intermediaires
            if not is_virtual_start and not(virtual_end and nnode == 0):
                distances.append(float(dist)/1000)
                durees.append(f"{dur//3600} h {(dur%3600)//60} min {(dur%3600)%60} secondes")


            index = nxt

        # Fin de tournée
        if not virtual_end:
            seq.append("Établissement")
            prenoms.append(etablissement.code)
            noms.append(etablissement.nom)
            adresses.append(etablissement.adresse.num_et_rue)
            villes.append(etablissement.adresse.ville)


        if seq != ["Établissement","Établissement"] and seq != ["Établissement"]:
            routes.append({
                "nom_tournee": groupe.nom,
                "vehicle_id": vid,
                "sequence": seq,
                "noms": noms,
                "prenoms": prenoms,
                "adresses": adresses,
                "villes": villes,
                "distances": distances,
                "durees": durees,
                "distance_totale": f"{float(dist_tot)/1000} km",
                "duree_totale": f"{dur_tot//3600} h {(dur_tot%3600)//60} min {(dur_tot%3600)%60} secondes"

            })

    return {"status": "SUCCESS", "routes": routes}

def solve_vrp(groupe, etablissement, capacities, time_limit, calculation_mod, mode="CLOSED"):
    print("Mode reçu :", mode)
    
    # Levée d'exception
    if not etablissement.adresse or not etablissement.adresse.latitude or not etablissement.adresse.longitude:
        raise ValueError("L'établissement n'a pas de coordonnées valides.")

    for enfant in groupe.enfants.all():
        if not enfant.adresse or not enfant.adresse.latitude or not enfant.adresse.longitude:
            raise ValueError(f"L'enfant {enfant} n'a pas de coordonnées géographiques.")
    
    # On crée la liste en ajoutant le dépot mais on attend les éventuelles options pour finaliser 
    coord_list = [
        f"{c.adresse.longitude},{c.adresse.latitude}"
        for c in groupe.enfants.all()
    ]
    depot = f"{etablissement.adresse.longitude},{etablissement.adresse.latitude}"
    coords_list = [depot] + coord_list

    # Ajout de dépot virtuel au début où à la fin en fonction des options.
    virtual_start = False
    virtual_end = False

    if mode == "NO_START":
        virtual_start = True
    elif mode == "NO_END":
        virtual_end = True

    coords = ";".join(coords_list) # Termine le formatage de la string qu'on va passer en paramètre à l'API OSRM
    distMat, durMat = get_matrices(coords) # Appel de l'API

    # Adaptation de la matrice si dépôt virtuel
    if virtual_start:
        # Ligne en plus au début avec zéros vers tous les noeuds (sauf soi-même) pour pouvoir accéder facilement à tous les noeuds
        row = [0] * (len(distMat[0]) + 1) # [0, 0, 0, ..., 0]
        distMat.insert(0, row) # Insère la ligne au début (à l'index 0)
        durMat.insert(0, row.copy()) # Ici le .copy() permet une séparation en mémoire de la première ligne sans quoi on aurait un problème d'accès
        for r in distMat[1:]: # Insère une colonne avec des valeurs très grandes hormis la première ligne (On évite ainsi les retours au dépôt)
            r.insert(0, 999999)  # Interdire retour vers dépôt de départ virtuel
        for r in durMat[1:]:
            r.insert(0, 999999)
        depot_ix = 0
        offset = 1 # Décalage des indices de ma matrice à cause de l'ajout de ma ligne et colonne
    elif virtual_end:
        # Ligne en plus au début avec un chiffre très grand vers tous les nœuds (sauf soi-même) pour s'assurer que ce soit le dernier noeuds
        row = [999999] * (len(distMat[0]) + 1) # [999999, 999999, ..., 999999]
        distMat.insert(0, row) # Insère la ligne au début (à l'index 0)
        durMat.insert(0, row.copy()) # Ici le .copy() permet une séparation en mémoire de la première ligne sans quoi on aurait un problème d'accès
        for r in distMat[1:]: # Insère une colonne avec des valeurs très petites hormis la première ligne, on s'assure l'accès l'accès facile depuis n'importe quel autre point.
            r.insert(0, 0)  # Faciliter retour vers dépôt virtuel
        for r in durMat[1:]:
            r.insert(0, 0)
        depot_ix = 0
        offset = 1 # Décalage des indices de ma matrice à cause de l'ajout de ma ligne et colonne
    else:
        depot_ix = 0
        offset = 0

    # Définition des variables
    nb_nodes = len(distMat)
    num_veh  = len(capacities)

    # Définition des starts/ends
    if virtual_start:
        starts = [depot_ix] * num_veh # indice du dépôt virtuel
        ends = [depot_ix + offset] * num_veh # indice de l'établissement
    elif virtual_end:
        starts = [depot_ix + offset] * num_veh # indice de l'établissement
        ends = [depot_ix] * num_veh # indice du dépôt virtuel
    else:
        starts = [depot_ix] * num_veh # comme pas d'ajout de ligne => indice de l'établissement
        ends = [depot_ix] * num_veh # idem
    
    # Création de la table d'adressage
    manager = pywrapcp.RoutingIndexManager(
        nb_nodes, num_veh, starts, ends
    )

    # Création du modèle
    routing = pywrapcp.RoutingModel(manager)

    # Définition de la fonction de callback des coûts
    def cost_cb(i, j):
        ni = manager.IndexToNode(i)
        nj = manager.IndexToNode(j)
        try:
            return distMat[ni][nj] if calculation_mod == "Distance" else durMat[ni][nj]
        except IndexError as e:
            print(f"[Erreur INDEX] i={i}, j={j}, ni={ni}, nj={nj}")
            print(f"Dimensions distMat: {len(distMat)}x{len(distMat[0])}")
            raise e  # Rejette quand même pour que tu vois l’erreur dans la console
    
    tc_index = routing.RegisterTransitCallback(cost_cb) # Enregistrement de la fonction de callback des coûts dans le modèle
    routing.SetArcCostEvaluatorOfAllVehicles(tc_index) # On dit au solver que c'est cette fonction qu'il doit utiliser quand il cherche le coût d'un arc

    # Définition de la fonction de callback des capacités
    def capacity_cb(i):
        ni = manager.IndexToNode(i)
    
        # Interdire de visiter le dépôt réel comme point client
        if (virtual_start or virtual_end) and ni == depot_ix + offset:
            return 0  # Pas de charge car c'est l'établissement
        elif ni == depot_ix:
            return 0  # Noeud virtuel ou etablissement dans le cas closed 
        else:
            return 1  # Tous les autres sont des clients


    dc_index = routing.RegisterUnaryTransitCallback(capacity_cb) # Enregistrement de la fonction de callback des capacités dans le modèle
    routing.AddDimensionWithVehicleCapacity( # On dit au solver que c'est cette fonction qu'il doit utiliser quand il cherche la capacité d'un noeud
        dc_index, # la fonction de callback des capacités
        0, # aucune tolèrance (capacité supplémentaire autorisée)
        capacities, # liste des capacités max pour chaque véhicule
        True, # atteste que la charge au départ est nulle
        "capacity" # nom de la dimension concernée
    )

    # Défintion des paramètres de recherche
    search_params = pywrapcp.DefaultRoutingSearchParameters()
    search_params.first_solution_strategy = (
        routing_enums_pb2.FirstSolutionStrategy.PATH_MOST_CONSTRAINED_ARC
    ) # Ici on choisit l'heuristique de recherche pour une première solution trouvée rapidement qui sera par la suite optimisée.
    # AUTOMATIC = Laisse OR-Tools choisir | PATH_CHEAPEST_ARC = Prends l'arc le moins cher possible
    # SAVINGS = Heuristique de Clarke-Wright (raisonnement en termes d'économies) | CHRISTOFIDES = Heuristique spécifique au TSP
    # PATH_MOST_CONSTRAINED_ARC = Prends l'arc le plus contraint (celui qui a le moins de voisins valides restants) | ALL_UNPERFORMED = ne fait rien, utile pour du debug
    
    search_params.local_search_metaheuristic = (
        routing_enums_pb2.LocalSearchMetaheuristic.GUIDED_LOCAL_SEARCH
    ) # Là on choisit l'heuristique de recherche d'une deuxième solution à partir de la première trouvée
    # AUTOMATIC = Laisse OR-Tools choisir | GUIDED_LOCAL_SEARCH = Ajoute des penalités aux arcs utilisés trop souvent pour trouver de nouvelles solutions
    # TABU_SEARCH = Ne revient pas sur les arcs déjà utilisés | SIMULATED_ANNEALING = Acceptation contrôlée de mauvaises solutions
    # GREEDY_DESCENT = Améliore vite (algo glouton) mais peux vite se bloquer
    
    search_params.time_limit.seconds = time_limit # Temps de recherche autorisé

    # Lancement de la résolution
    sol = routing.SolveWithParameters(search_params)
    if not sol:
        return {"status": "FAILURE", "routes": []}

    # Si une solution est trouvée, on formate
    return format_solution(manager, routing, sol, distMat, durMat, groupe, etablissement, virtual_start, virtual_end)

def nettoyage_nom_excel(nom_brut):
    nom_sans_accent = ''.join(
        c for c in unicodedata.normalize('NFKD', nom_brut)
        if not unicodedata.combining(c)
    )

    # Remplacer les caractères interdits par un underscore
    nom_nettoye = re.sub(r'[^A-Za-z0-9 _-]', '_', nom_sans_accent)

    # Supprimer les espaces en début/fin
    nom_nettoye = nom_nettoye.strip()

    # Limiter la longueur 
    nom_nettoye = nom_nettoye[:50]  # exemple: on limite à 50 caractères

    return nom_nettoye

def export_to_excel_formatted(data):
    wb = Workbook() # Création d'un fichier excel 
    ws = wb.active # On stipule sur quelle feuille de travail on se place
    ws.title = nettoyage_nom_excel(data["routes"][0]["nom_tournee"]) # Ajout d'un titre

    for i, route in enumerate(data["routes"]):
        if i != 0:
            ws.append([]) # lignes vides entre les blocs
            ws.append([]) 
            

        # En-têtes colonnes
        nb_etapes = len(route['sequence'])
        cols = ["Départ"] + [f"Enfant {i}" for i in range(1, nb_etapes - 1)] + ["Fin"]
        ws.append([f"Taxi {i+1}"] + cols)

        # Initialiser les lignes
        lignes = {
            "Nom": [],
            "Prénom": [],
            "Adresse": [],
            "Ville": [],
            "Distance": [],
            "Temps": [],
        }

        for idx in range(nb_etapes):
            lignes["Nom"].append(route["noms"][idx])
            lignes["Prénom"].append(route["prenoms"][idx])
            lignes["Adresse"].append(route["adresses"][idx])
            lignes["Ville"].append(route["villes"][idx])
            if idx == 0:
                lignes["Distance"].append("0.0 km")
                lignes["Temps"].append("0 h 0 min 0 secondes")
            else:
                lignes["Distance"].append(f"{route["distances"][idx - 1]} km")
                lignes["Temps"].append(route["durees"][idx - 1])

        # Écrire dans le fichier
        for key in ["Nom", "Prénom", "Adresse", "Ville", "Distance", "Temps"]:
            ligne = [key] + lignes[key]
            ws.append(ligne)

        # Footer distance totale et duree totale
        ws.append([""])
        ws.append(["Distance totale"] + [route["distance_totale"]] + ["Temps total"] + [route["duree_totale"]])

    # Style : mettre "Taxi X" en gras
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if isinstance(cell.value, str):
                if cell.value.startswith("Taxi"):
                    cell.font = Font(bold=True, size=12)
                elif cell.value in ("Distance totale", "Temps total"):
                    cell.font = Font(bold=True, size=12)
                else:
                    if cell.column == 1 or (isinstance(row[0].value, str) and row[0].value.startswith("Taxi")):
                        cell.font = Font(italic=True)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for col_cells in ws.columns:
        # id de colonne (A, B, C…)
        col_letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for c in col_cells:
            if c.value is not None:
                # longueur du texte (sans styles)
                max_len = max(max_len, len(str(c.value)))
        # marge + plafond pour éviter des colonnes démesurées
        ws.column_dimensions[col_letter].width = max(14, min(45, max_len + 2))

    # Sauvegarde dans un buffer en mémoire
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


